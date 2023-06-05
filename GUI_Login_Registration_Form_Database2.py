import os
import smtplib
import ssl
from datetime import date, datetime, timedelta

import pandas as pd
from email_validator import validate_email, EmailNotValidError
from openpyxl import *
from tkinter import *
from twilio.rest import Client



# open existing excel files for user infro
# filesshould be saved in the same directory ot pathrrrrr
wb = load_workbook(".\SimpleRegistrationDatabase.xlsx")
wb2 = load_workbook(".\PasswordUser.xlsx")



class ExcelSheetFormatter:
    def __init__(self, sheet):
        self.sheet = sheet
        self.column_widths = {
            "A": 35,
            "B": 35,
            "C": 35,
            "D": 20,
            "E": 20,
            "F": 40,
            "G": 50,
            "H": 20,
            "I": 20,
            "J": 20,
            "K": 20,
        }
        self.headers = [
            "FirstName",
            "MiddleName",
            "LastName",
            "DOB",
            "PhoneNumber",
            "Email",
            "StreetAddress",
            "CurrentDate",
            "FutureDate",
            "Sent",
            "SendText",
        ]

    def format_sheet(self):
        # Set column widths
        for column_letter, width in self.column_widths.items():
            self.sheet.column_dimensions[column_letter].width = width

        # Set headers
        for index, header in enumerate(self.headers, start=1):
            self.sheet.cell(row=1, column=index).value = header

# Create sheets for objects
sheet = wb.active
sheet2 = wb2.active

# Create formatter and format sheet
formatter = ExcelSheetFormatter(sheet)


# set focus on the middle_field box
def focus1(event):
    middle_field.focus_set()

def focus2(event):
    last_field.focus_set()

def focus3(event):
    DOB_field.focus_set()

def focus4(event):
    phone_field.focus_set()

def focus5(event):
    email_field.focus_set()

def focus6(event):
    address_field.focus_set()

def focus7(event):
    current_date_field.focus_set()


# call function to clear text entry boxes
def clear():

    # clear the content of text entry box
    name_field.delete(0, END)
    middle_field.delete(0, END)
    last_field.delete(0, END)
    DOB_field.delete(0, END)
    phone_field.delete(0, END)
    email_field.delete(0, END)
    address_field.delete(0, END)
    current_date_field.delete(0, END)


# Empty input screen to notify user of empty text boxes
# If any one text box is empty is empty the form will not submit and
# the user wilbe notifies with a pop up screen
def empty_input_error():
    global empty_input_screen
    empty_input_screen.destroy()


def Empty_Input():
    global empty_input_screen
    empty_input_screen = Toplevel()
    empty_input_screen.title("Error")
    empty_input_screen.geometry("200x100")
    Label(empty_input_screen, text="Empty text entry!", font=("Calibri", 13)).grid(row=0, column=0, padx=25)
    Button(empty_input_screen, text="OK to exit", width=15, height=1, bg="red", font=("Calibri", 13), command=empty_input_error).grid(row=1, column=0, padx=25)



# Notify the user of a susseccfull registration form submition_success
# Close out registration form upon pressing ok to exit button
def delete_submition_success_screen():
    global submition_success_screen
    submition_success_screen.destroy()


def submition_success():
    global submition_success_screen
    submition_success_screen = Toplevel()
    submition_success_screen.title("Submission Success")
    submition_success_screen.geometry("200x100")
    Label(submition_success_screen, text="Submission Success!", font=("Calibri", 13)).grid(
        row=0, column=0, padx=25)
    Button(submition_success_screen, text="OK to exit", width=15, height=1, bg="red", font=("Calibri", 13), command=delete_submition_success_screen).grid(row=1, column=0, padx=25)



# Verify that a second email will be sent to recipient that just registered
# 21 days from the first vaccine date
def send_email(from_address, password, to_address, subject, message):
    # Validate the email address
    try:
        validate_email(to_address, check_deliverability=False)
    except EmailNotValidError as e:
        print(f"Invalid email address: {to_address}. {str(e)}")
        return

    # If the email address is valid, send the email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(from_address, password)
        server.sendmail(from_address, to_address, f"Subject: {subject}\n\n{message}")


def send_text_message(from_number, to_number, message):
    client = Client("EnteryourTwilioSSID", "EnteryourTwilioAuthToken")
    try:
        message = client.messages.create(body=message, from_=from_number, to=to_number)
        return message.sid
    except Exception as e:
        print(f"Failed to send text message: {e}")

def first_vaccine():
    df = pd.read_csv("COVID_Database.csv")
    rows = df.shape[0]

    first_vac_date = datetime.strptime(df.iloc[rows - 1, 7], "%Y-%m-%d %H:%M:%S").strftime("%B %d, %Y")
    sec_vac_date = datetime.strptime(df.iloc[rows - 1, 8], "%Y-%m-%d %H:%M:%S").strftime("%B %d, %Y")

    first_name = df.iloc[rows - 1, 0]
    last_name = df.iloc[rows - 1, 2]
    email = df.iloc[rows - 1, 5]
    phone_number = df.iloc[rows - 1, 10]

    email_subject = "First COVID-19 Vaccination Received"
    email_message = f"Hello {first_name} {last_name}, you received your first COVID-19 Vaccination today on {first_vac_date}.\n\nYour second dose for your complete Vaccination is on {sec_vac_date}."

    email_from_address = "EnterYourEmailHere"
    email_password = "EnterYourPasswordHere"

    text_message = email_message

    send_email(email_from_address, email_password, email, email_subject, email_message)
    send_text_message("EnterTwilioPhoneNumberHere", phone_number, text_message)



# Second vaccine notification function
def second_vaccine():
    df2 = pd.read_csv("COVID_Database.csv")
    
    message_template = """Hello {First} {Last}, your second COVID-19 vaccination is coming up on {Vaccination2}"""
    check_date = datetime.now() - timedelta(days=3)

    from_address = "EnterYourEmailHere"
    password = "Enetryourpasswordhere"
    from_number = "EnterTwilioPhoneNumberHere"

    for i, r in df2.iterrows():
        future_date = datetime.strptime(r["FutureDate"], "%Y-%m-%d %H:%M:%S")
        if future_date <= check_date and r["Sent"] == 1:
            message = message_template.format(
                First=r["FirstName"],
                Last=r["LastName"],
                Vaccination2=future_date.strftime("%B %d, %Y"))
            
            # send the email
            send_email(from_address, password, r["Email"], "Reminder for Second COVID-19 Vaccination", message)

            # send the text message
            send_text_message(from_number, r["SendText"], message)

            df2.loc[i, "Sent"] = 0
            df2.to_excel("SimpleRegistrationDatabase.xlsx", index=False)
            df2.to_csv("COVID_Database.csv", index=None, header=True)

# Write unser input into excel sheet2
# if either text entry box is empty the user will be notified
def insert():
    if (
        name_field.get() == ""
        or last_field.get() == ""
        or DOB_field.get() == ""
        or phone_field.get() == ""
        or email_field.get() == ""
        or address_field.get() == ""
        or current_date_field.get() == ""
    ):
        Empty_Input()
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = middle_field.get()
        sheet.cell(row=current_row + 1, column=3).value = last_field.get()
        sheet.cell(row=current_row + 1, column=4).value = DOB_field.get()
        sheet.cell(row=current_row + 1, column=5).value = phone_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        sheet.cell(row=current_row + 1, column=8).value = datetime.strptime(current_date_field.get(), "%Y-%m-%d") + timedelta(minutes=5)

        sheet.cell(row=current_row + 1, column=9).value = datetime.strptime(current_date_field.get(), "%Y-%m-%d") + timedelta(days=21)
        sheet.cell(row=current_row + 1, column=10).value = 1
        sheet.cell(row=current_row + 1, column=11).value = "+1" + phone_field.get()

        wb.save(".\SimpleRegistrationDatabase.xlsx")

        df_csv = pd.read_excel("SimpleRegistrationDatabase.xlsx")
        df_csv.to_csv("COVID_Database.csv", index=None, header=True)

        name_field.focus_set()
        clear()
        submition_success()
        first_vaccine()
        second_vaccine()


# impliment user registry form window for user registration
def register_form():
    global root
    root = Toplevel()
    root.title("Registration form")
    root.geometry("1000x400")

    bg = PhotoImage(file="black2.png", height=400, width=1000)
    bg_label = Label(root, image=bg)
    bg_label.image = bg
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    global Name_in, course_in, sem_in, grad_in, phone_in, email_in, address_in, current_date_in

    global name_field, middle_field, last_field, DOB_field, phone_field, email_field, address_field, current_date_field

    input_vars = {
        "Name": StringVar(),
        "Course": StringVar(),
        "Sem": StringVar(),
        "Grad": StringVar(),
        "Phone": StringVar(),
        "Email": StringVar(),
        "Address": StringVar(),
        "Current Date": StringVar()
    }

    formatter.format_sheet()

    labels = {
        "name": "First Name",
        "course": "Middle Name",
        "sem": "Last Name",
        "grad": "DOB",
        "phone": "10 digit Phone No. (0123456789)",
        "email": "Email",
        "address": "Street Address",
        "current_date": "Current Date (YYYY-MM-DD)",
    }

    for row, key in enumerate(labels.keys(), start=1):
        label = Label(root, text=labels[key], bg="black", fg="white", font=("Calibri", 13))
        label.grid(row=row, column=0)

    name_field = Entry(root, textvariable=input_vars["Name"])
    middle_field = Entry(root, textvariable=input_vars["Course"])
    last_field = Entry(root, textvariable=input_vars["Sem"])
    DOB_field = Entry(root, textvariable=input_vars["Grad"])
    phone_field = Entry(root, textvariable=input_vars["Phone"])
    email_field = Entry(root, textvariable=input_vars["Email"])
    address_field = Entry(root, textvariable=input_vars["Address"])
    current_date_field = Entry(root, textvariable=input_vars["Current Date"])

    name_field.bind("<Return>", focus1)
    middle_field.bind("<Return>", focus2)
    last_field.bind("<Return>", focus3)
    DOB_field.bind("<Return>", focus4)
    phone_field.bind("<Return>", focus5)
    email_field.bind("<Return>", focus6)
    address_field.bind("<Return>", focus7)

    name_field.grid(row=1, column=1, ipadx="100")
    middle_field.grid(row=2, column=1, ipadx="100")
    last_field.grid(row=3, column=1, ipadx="100")
    DOB_field.grid(row=4, column=1, ipadx="100")
    phone_field.grid(row=5, column=1, ipadx="100")
    email_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")
    current_date_field.grid(row=8, column=1, ipadx="100")

    submit = Button(root, text="Submit", fg="Black", relief="flat", bg="grey", width=15, height=1, font=("Calibri", 13), command=insert)
    submit.grid(row=9, column=1)


# design register screen
def register():
    global register_screen
    register_screen = Toplevel(main_screen)
    register_screen.title("Register")
    register_screen.geometry("300x420")

    # Background image for window
    # refrenced above
    bg = PhotoImage(file="black2.png", height=420, width=300)
    bg_label = Label(register_screen, image=bg)
    bg_label.image = bg
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    global username, username_entry
    username = StringVar()
    global password, password_entry
    password = StringVar()

    # Label displayed for users to submit username and password information
    Label(register_screen, text="Please enter details below to register", fg="white", bg="black", width="30", height="2", font=("Calibri", 13)).grid(row=0, column=0, padx="10")

    # Labels for user entries and text entries
    # Buttons changes include font color and background color
    username_label = Label(register_screen, text="Username * ", fg="white", bg="black", font=("Calibri", 13))
    username_label.grid(row=2, column=0)
    username_entry = Entry(register_screen, textvariable=username)
    username_entry.grid(row=3, column=0)
    password_label = Label(register_screen, text="Password * ", fg="white", bg="black", font=("Calibri", 13))
    password_label.grid(row=4, column=0)
    password_entry = Entry(register_screen, textvariable=password, show="*")
    password_entry.grid(row=5, column=0)

    Button(register_screen, text="Register", width=15, height=1, bg="grey", font=("Calibri", 13), command=register_user).grid(row=7, column=0)


# Designing window for login
def login():
    global login_screen, username_verify, password_verify, username_login_entry, password_login_entry
    login_screen = Toplevel(main_screen)
    login_screen.title("Login")
    login_screen.geometry("320x400")

    bg = PhotoImage(file="black2.png", height=400, width=320)
    bg_label = Label(login_screen, image=bg)
    bg_label.image = bg
    bg_label.grid(row=0, column=0, columnspan=20, rowspan=20)

    Label(login_screen, text="Please enter details below to login", fg="white", bg="black", width="30", height="2", font=("Calibri", 13)).grid(row=0, column=0, padx="10")

    username_verify = StringVar()
    password_verify = StringVar()

    Label(login_screen, text="Username * ", fg="white", bg="black", font=("Calibri", 13)).grid(row=3, column=0)
    username_login_entry = Entry(login_screen, textvariable=username_verify)
    username_login_entry.grid(row=4, column=0)

    Label(login_screen, text="Password * ", fg="white", bg="black", font=("Calibri", 13)).grid(row=6, column=0)
    password_login_entry = Entry(login_screen, textvariable=password_verify, show="*")
    password_login_entry.grid(row=7, column=0)

    Button(login_screen, text="Login", width=15, height=1, bg="grey", font=("Calibri", 13), command=login_verify).grid(row=9, column=0)



# write user name and password to spreadhseet
def excel2():
    # Resize the width of columns in the excel spreadsheet
    sheet2.column_dimensions["A"].width = 30
    sheet2.column_dimensions["B"].width = 10

    # Write given data to an excel spreadsheet at particular location
    # Columns for username and password
    column_labels = ["User Name", "Password"]
    for col, label in enumerate(column_labels, start=1):
        sheet2.cell(row=1, column=col).value = label



# Implementing event on register button
def register_user():
    username_info = username.get()
    password_info = password.get()
    excel2()

    # Get the current row and column in the spreadsheet
    current_row2 = sheet2.max_row
    current_column2 = sheet2.max_column

    # Write the username and password information to the spreadsheet
    sheet2.cell(row=current_row2 + 1, column=1).value = username_info
    sheet2.cell(row=current_row2 + 1, column=2).value = password_info

    wb2.save(".\PasswordUser.xlsx")

    # Clear the entry fields
    username_entry.delete(0, END)
    password_entry.delete(0, END)

    # Display registration success message
    Label(register_screen, text="Registration Success", bg="black", fg="red", font=("calibri", 11)).grid(row=10, column=0)


# Implementing event on login button
def login_verify():
    username1 = username_verify.get()
    password1 = password_verify.get()
    username_login_entry.delete(0, END)
    password_login_entry.delete(0, END)

    names = []
    passwords = []

    # Populate the names and passwords lists from the spreadsheet
    for i in range(2, sheet2.max_row + 1):
        if sheet2.cell(row=i, column=1).value:
            names.append(sheet2.cell(row=i, column=1).value)
            passwords.append(sheet2.cell(row=i, column=2).value)

    try:
        index = names.index(username1)
    except ValueError:
        print("")

    if username1 in names:
        if password1 == passwords[index]:
            login_success()
            register_form()
        else:
            password_not_recognized()
    else:
        user_not_found()



# Designing popup for login succes
def login_success():
    global login_success_screen
    login_success_screen = Toplevel(login_screen)
    login_success_screen.title("Success")
    login_success_screen.geometry("200x100")

    # display Login success to user
    Label(login_success_screen, text="Login Success!", font=("Calibri", 13)).grid(row=0, column=0, padx=25)

    # close login success window when pressing button
    Button(login_success_screen, text="OK to exit", width=15, height=1, bg="red", font=("Calibri", 13), command=delete_login_success).grid(row=1, column=0, padx=25)


# Designing popup for login invalid password
def password_not_recognized():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(login_screen)
    password_not_recog_screen.title("Error")
    password_not_recog_screen.geometry("200x100")

    # notify user of invalid password
    Label(password_not_recog_screen, text="Invalid Password ", font=("Calibri", 13)).grid(row=0, column=0, padx=25)
    Button(password_not_recog_screen, text="OK to exit", width=15, height=1, bg="red", font=("Calibri", 13), command=delete_password_not_recognised).grid(row=1, column=0, padx=25)


# Designing popup for user not found
def user_not_found():
    global user_not_found_screen
    user_not_found_screen = Toplevel(login_screen)
    user_not_found_screen.title("Error")
    user_not_found_screen.geometry("200x100")

    # notify user of invalid username
    Label(user_not_found_screen, text="User Not Found", font=("Calibri", 13)).grid(row=0, column=0)
    Button(user_not_found_screen, text="OK to exit", width=15, height=1, bg="red", font=("Calibri", 13), command=delete_user_not_found_screen).grid(row=1, column=0, padx=25)


# Deleting popups
# Delete login success screen
def delete_login_success():
    login_success_screen.destroy()
    login_screen.destroy()


# Delete password not recognized screen
def delete_password_not_recognised():
    password_not_recog_screen.destroy()


# Delete uset not found screen
def delete_user_not_found_screen():
    user_not_found_screen.destroy()


# delete login screen
def delete_login_screen():
    login_screen.destroy()


# Designing Main(first) window
def main_account_screen():
    # Create the main screen
    global main_screen
    main_screen = Tk()
    main_screen.geometry("800x400")
    main_screen.title("Immunization")

    # Set the background image
    bg = PhotoImage(file="black2.png", height=400, width=800)
    bg_label = Label(main_screen, image=bg)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    # Create the main text label
    main_label = Label(main_screen, text="Set Up Vaccination Appointment", fg="white", font=("Calibri", 20), bg="black")
    main_label.place(x=50, y=90, width=700)

    # Create the login and register buttons
    Button(text="Login", height="2", width="30", bg="grey", command=login).place(x=280, y=200)
    Button(text="Register", height="2", width="30", bg="grey", command=register).place(x=280, y=250)

    # Start the main event loop
    main_screen.mainloop()

main_account_screen()
